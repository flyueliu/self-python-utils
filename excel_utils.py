import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class ExcelEditor(object):

    def __init__(self, file_name: str):
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook.active

    def get_workbook(self) -> Workbook:
        return self.workbook

    def change_sheet(self, index: int) -> Worksheet:
        self.sheet = self.workbook[self.workbook.worksheets[index]]
        return self.sheet

    def get_cell(self, row: int = None, column: int = None, name: str = None) -> Cell:
        if (row is not None) and (column is not None):
            return self.sheet.cell(row, column)
        if name:
            return self.sheet[name]
        raise TypeError("input argument error!")

    """
          获取单元格数据
          row:行号，从1开始
          column:列号，从1开始
    """

    def get_cell_value(self, row: int = None, column: int = None, name=None):
        return self.get_cell(row=row, column=column, name=name).value

    """
      设置单元格数据
      row:行号，从1开始
      column:列号，从1开始
      name: 
    """

    def set_cell(self, row: int = None, column: int = None, name: str = None, value=None, style: str = None):
        if (row is not None) and (column is not None):
            edit_cell = self.sheet.cell(row, column, value=value)
        elif name:
            edit_cell = self.sheet[name]
            edit_cell.value = value
        else:
            raise ValueError("unknown cell position ")
        if style:
            edit_cell.number_format = style

    def get_max_row(self) -> int:
        return self.sheet.max_row

    def get_max_column(self) -> int:
        return self.sheet.max_column

    def save(self, name):
        self.workbook.save(name)
        self.workbook.close()

    def __del__(self):
        self.workbook.close()


if __name__ == '__main__':
    excel = ExcelEditor(r'others/test.xlsx')
    for i in range(1, excel.get_max_row() + 1):
        for j in range(1, excel.get_max_column() + 1):
            value = excel.get_cell_value(i, j)
            print(value, end="|")
        print(end='\n')
    excel.save("user.xlsx")
